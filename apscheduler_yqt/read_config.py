# -*- coding: utf-8 -*-
"""
   File Name：     read_config
   Description :
   Author :       hike
   time：          2021/3/26 10:05
"""

from openpyxl import load_workbook

#数据加载
wb=load_workbook('config.xlsx',data_only=True)
sheet=wb['Sheet1']
row_info=[]
for row in sheet.iter_rows(min_row=2,max_col=5,max_row=2):
    for cell in row:
        row_info.append(cell.value)

pinyin=['yuqingtong_username','yuqingtong_password','dama_username','dama_password','softid']
info=dict(list(zip(pinyin,row_info)))

USERNAME = info['yuqingtong_username']
PASSWORD = info['yuqingtong_password']

time_data_all=sheet.iter_rows(min_row=7)
times=[]
for row in time_data_all:
    time = {
        'start_time':'',
        'end_time':'',
        'time_delay':''
    }
    for cell in row:
        if(cell.column==1):
            time['start_time']=cell.value
        elif(cell.column==2):
            time['end_time']=cell.value
        elif(cell.column==3):
            time['time_delay']=cell.value
    times.append(time)
