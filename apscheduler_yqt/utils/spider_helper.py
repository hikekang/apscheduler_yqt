"""
爬虫帮助
    1.验证码识别
    2.文档保存
"""
import requests
import json
import csv
import pandas as pd
import os
from yuqingtong.config import *
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class SpiderHelper(object):

    @staticmethod
    def recognise_code(image_base64,infos):
        # base64_data = base64.b64encode(img_bytes)
        # img_b64_data = base64_data.decode()
        # data = {
        #     "user": info['dama_username'],
        #     "pass": info['dama_password'],
        #     "softid": info['softid'],
        #     "codetype": 1006,
        #     "file_base64": image_base64
        # }
        info=infos.getDictBySection('chaojiyin_info')
        data = {
            "user": info['username'],
            "pass": info['pwd'],
            "softid": info['softid'],
            "codetype": 1006,
            "file_base64": image_base64
        }
        headers = {"Content-Type": "application/json"}
        try:
            proxies = {"http": None, "https": None}
            response = requests.post(url="http://upload.chaojiying.net/Upload/Processing.php", data=json.dumps(data),
                                     headers=headers, proxies=proxies)
            print(response.json())
            return response.json().get("pic_str")
        except Exception as e:

            print(e)

    @staticmethod
    def save_csv(data_list, out_file):
        # print(out_file)
        dir = os.path.dirname(out_file)
        # print(dir)
        if not os.path.exists(dir):
            os.mkdir(dir)
        # 利用datafram 获取列索引
        df = pd.DataFrame(data_list)
        # print(list(df.columns))

        # 通过判断时候有csv文件，来判断是否写表头，避免重复写表头
        is_writeheader = not os.path.exists(out_file)
        with open(out_file, 'a', encoding="utf-8-sig", newline="") as csvfile:

            writer = csv.DictWriter(csvfile, fieldnames=df.columns)
            if is_writeheader:

                writer.writeheader()  # 写入表头

            for data in data_list:
                writer.writerow(data)
    @staticmethod
    def save_xlsx(data_list, out_file,sheet_name):
        # threadLock = threading.Lock()
        # threadLock.acquire()
        head_xlsx = ['时间', '标题', '描述', '链接', '转发内容', '发布人', 'attitude', 'images', 'reposts_count', 'comments_count',
                     'sort', 'industry', 'related_words', 'site_name', 'area']

        dir = os.path.dirname(out_file)
        if not os.path.exists(dir):
            os.makedirs(dir)
        if not os.path.exists(out_file):
            wb=Workbook(out_file)
            sheet=wb.create_sheet(title=sheet_name)
            #写入表头
            # for i,item in enumerate(head_xlsx):
            #     sheet.cell(row=1,column=i+1,value=item)
            sheet.append(head_xlsx)
            for data in data_list:
                values=(data[k] for k in head_xlsx)
                sheet.append(values)
            wb.save(out_file)
            wb.close()
        else:
            wb=load_workbook(out_file)
            sheet=wb[sheet_name]
            for data in data_list:
                values = (data[k] for k in head_xlsx)
                sheet.append(values)
            wb.save(out_file)
            wb.close()
        # threadLock.release()
    @staticmethod
    def save_record(out_file,yq_number,xlsx_num,post_num,post_num2,sql_num,data_list):
        '''
        创建一个路径+文件
        拿到
            ①拿到舆情通数量
            ②xlsx 数量
            ③post返回数量
            ④seq入库查询数量
        记录数值
        计算差值 70%以上
            数据量相差太多预警并且记录
        '''
        head_xlsx=['项目名称','抓取时间','开始时间','结束时间','舆情通数量','xlsx数量','post数量','seq数量','最终差异']
        dir = os.path.dirname(out_file)
        if not os.path.exists(dir):
            os.makedirs(dir)
        if not os.path.exists(out_file):
            wb = Workbook(out_file)
            sheet = wb.create_sheet(title='record')
            # 写入表头
            # for i,item in enumerate(head_xlsx):
            #     sheet.cell(row=1,column=i+1,value=item)
            sheet.append(head_xlsx)
            data_list.append(yq_number)
            data_list.append(xlsx_num)
            data_list.append(post_num)
            data_list.append(post_num2)
            data_list.append(sql_num)
            chayi=yq_number-sql_num
            if chayi/yq_number>=0.3:
                print('预警')
            data_list.append(chayi)
            sheet.append(data_list)
            wb.save(out_file)
            wb.close()
            print("记录完成")
        else:
            wb = load_workbook(out_file)
            sheet = wb['record']
            data_list.append(yq_number)
            data_list.append(xlsx_num)
            data_list.append(post_num)
            data_list.append(post_num2)
            data_list.append(sql_num)
            chayi = yq_number - sql_num
            if chayi/yq_number>=0.3:
                print('预警')
            data_list.append(chayi)
            sheet.append(data_list)
            wb.save(out_file)
            wb.close()
            print("记录完成")

    @staticmethod
    def save_record_auto(out_file, yq_number,post_number,sql_num_A,sql_num_B, data_list):
        '''
        创建一个路径+文件
        拿到
            ①拿到舆情通数量
            ②xlsx 数量
            ③post返回数量
            ④seq入库查询数量
        记录数值
        计算差值 70%以上
            数据量相差太多预警并且记录
        '''
        # head_xlsx = ['项目名称', '抓取时间', '开始时间', '结束时间', '舆情通数量', 'xlsx数量', 'post数量', 'seq数量', '最终差异']
        head_xlsx = ['项目名称', '抓取时间', '开始时间', '结束时间', '舆情通数量','上传数量', 'seqA数量', 'seqB数量','最终差异']
        dir = os.path.dirname(out_file)
        if not os.path.exists(dir):
            os.makedirs(dir)
        if not os.path.exists(out_file):
            wb = Workbook(out_file)
            sheet = wb.create_sheet(title='record')
            # 写入表头
            # for i,item in enumerate(head_xlsx):
            #     sheet.cell(row=1,column=i+1,value=item)
            sheet.append(head_xlsx)
            data_list.append(yq_number)
            data_list.append(post_number)
            data_list.append(sql_num_A)
            data_list.append(sql_num_B)
            chayi = yq_number - sql_num_A
            # if chayi / yq_number >= 0.3:
            #     print('预警')
            data_list.append(chayi)
            sheet.append(data_list)
            wb.save(out_file)
            wb.close()
            print("记录完成")
        else:
            wb = load_workbook(out_file)
            sheet = wb['record']
            data_list.append(yq_number)
            data_list.append(post_number)
            data_list.append(sql_num_A)
            data_list.append(sql_num_B)
            chayi = post_number - sql_num_A
            # if chayi / yq_number >= 0.3:
            #     print('预警')
            data_list.append(chayi)
            sheet.append(data_list)
            wb.save(out_file)
            wb.close()
            print("记录完成")

    @staticmethod
    def save_record_day_data(out_file, yq_number,sql_num_B):
        '''
        创建一个路径+文件
        拿到
            ①拿到舆情通数量
            ②xlsx 数量
            ③post返回数量
            ④seq入库查询数量
        记录数值
        '''
        data_list=[]
        head_xlsx = ['项目名称', '舆情通数量', '天颂数量']
        dir = os.path.dirname(out_file)
        if not os.path.exists(dir):
            os.makedirs(dir)
        if not os.path.exists(out_file):
            wb = Workbook(out_file)
            sheet = wb.create_sheet(title='报表')
            # 写入表头
            # for i,item in enumerate(head_xlsx):
            #     sheet.cell(row=1,column=i+1,value=item)
            sheet.append(head_xlsx)
            data_list.append(yq_number)
            data_list.append(sql_num_B)
        else:
            wb = load_workbook(out_file)
            sheet = wb['报表']
            data_list.append(yq_number)
            data_list.append(sql_num_B)
            sheet.append(data_list)
            wb.save(out_file)
            wb.close()
            print("记录完成")
    @staticmethod
    def all_project_save_record_day(out_file, yq_number,sql_num_B,project_name,industry_name):
        """
        :param out_file:文件路径
        :param yq_number:舆情通数量
        :param sql_num_B:系统数量
        :param project_name:项目名称
        :param industry_name:行业名称（sheet页名称）
        :return:
        """
        head_xlsx=['项目名称','舆情通数量', '天颂数量']
        all_industry_name_list=['流通贸易','金融业','IT业','房地产','汽车业','快消品','化妆品','旅游业']


        xlsx_dir=os.path.dirname(out_file)
        if not os.path.exists(xlsx_dir):
            os.makedirs(xlsx_dir)
        if not os.path.exists(out_file):
            wb=Workbook(out_file)
            for index, item in enumerate(all_industry_name_list):
                wb.create_sheet(item,index)

            for sheet in wb:
                # 自适应
                # for col in sheet.columns:
                #     max_length = 0
                #     column = col[0].column_letter  # Get the column name
                #     for cell in col:
                #         try:  # Necessary to avoid error on empty cells
                #             if len(str(cell.value)) > max_length:
                #                 max_length = len(str(cell.value))
                #         except:
                #             pass
                #     adjusted_width = (max_length + 2) * 1.2
                #     sheet.column_dimensions[column].width = adjusted_width

                for i in range(1,4):
                    # 设置列的宽度
                    sheet.column_dimensions[get_column_letter(i)].width=20
                sheet.append(head_xlsx)
                if sheet.title==industry_name:
                    sheet.append([project_name,yq_number,sql_num_B])
            wb.save(out_file)
            wb.close()
        else:
            wb=load_workbook(out_file)
            for sheet in wb:
                if sheet.title==industry_name:
                    sheet.append([project_name,yq_number,sql_num_B])
            wb.save(out_file)
            wb.close()


if __name__ == '__main__':
    # SpiderHelper.save_csv([{"name": 12}], os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "resle.csv"))
    # with open("./image.png", "rb") as f:
    #     img_bytes = f.read()
    # print(SpiderHelper.recognise_code(img_bytes))
    SpiderHelper.all_project_save_record_day("F\data\hike.xlsx",1,1,"优速","流通贸易")
    SpiderHelper.all_project_save_record_day("F\data\hike.xlsx",1,1,"优速","IT业")
