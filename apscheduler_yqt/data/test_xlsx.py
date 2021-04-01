# -*- coding: utf-8 -*-
"""
   File Name：     test_xlsx
   Description :
   Author :       hike
   time：          2021/3/31 11:59
"""

import os
from openpyxl import load_workbook
from openpyxl import Workbook

def save_xlsx(data_list, out_file):
    dir = os.path.dirname(out_file)
    if not os.path.exists(dir):
        os.mkdir(dir)

    wb = load_workbook(out_file)
    sheet = wb['Sheet1']
    # wb=Workbook(out_file)
    # sheet=wb.create_sheet('hike')
    # sheet=wb['sheet1']
    # sheet=wb.activate
    # 写入表头
    head_xlsx = ['时间', '标题']
    sheet.append(head_xlsx)
    # for i, item in data_list:
    #     sheet.cell(row=i + 2, column=i + 1, value=item)
    for data in data_list:
        values = (data[k] for k in head_xlsx)
        sheet.append(values)
    wb.save(out_file)
    wb.close()
datalist=[
    {'时间':'1','标题':'2'},
    {'时间':'1','标题':'2'},
    {'时间':'1','标题':'2'},
]
outfile=os.getcwd()
save_xlsx(datalist,'13.xlsx')