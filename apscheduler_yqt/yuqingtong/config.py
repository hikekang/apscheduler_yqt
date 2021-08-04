# -*- coding: utf-8 -*-
import configparser
import os
from apscheduler.triggers.cron import CronTrigger
from apscheduler.schedulers.blocking import BlockingScheduler

from openpyxl import load_workbook
#数据加载
# dir_path=os.path.dirname(os.path.dirname(os.path.relpath(__file__)))
# print(dir_path)
# config_xlsx_path=os.path.join(dir_path,'config.xlsx')
# print(config_xlsx_path)
# wb=load_workbook(config_xlsx_path,data_only=True)
# wb=load_workbook('../config.xlsx',data_only=True)
# dir_path=os.path.dirname(os.getcwd())
# config_xlsx_path = os.path.join(dir_path, r'config.xlsx')
# wb = load_workbook(config_xlsx_path, data_only=True)
# sheet=wb['info']
#
# # 获取用户名信息
# row_list=[]
# pinyin=['yuqingtong_username','yuqingtong_password','dama_username','dama_password','softid','project_name','sheet_name','keywords']
# for row in sheet.iter_rows(min_row=2,max_col=8):
#     row_info = []
#     for cell in row:
#         row_info.append(cell.value)
#     info = dict(list(zip(pinyin, row_info)))
#     row_list.append(info)


# for data in sheet[2:5]:
#     print(type(data))
#     print(data)
#     d=dict(list(zip(pinyin,data)))
#     print(d)
# print(row_info)

# print(info)
# USERNAME = info['yuqingtong_username']
# PASSWORD = info['yuqingtong_password']



# 获取设置的时间
def get_time_list():
    sheet=wb['time']
    time_data_all = sheet.iter_rows(min_row=2)
    times = []
    for row in time_data_all:
        time = {
            'start_time':'',
            'end_time':'',
            'time_delay':'2'
        }
        for cell in row:
            if(cell.column==1):
                time['start_time']=cell.value
            elif(cell.column==2):
                time['end_time']=cell.value
            elif(cell.column==3):
                time['time_delay']=cell.value
        times.append(time)
    return times

# USERNAME = info['yuqingtong_username']
# PASSWORD = info['yuqingtong_password']
# 浏览器相关
HEAD_LESS = False

# 等待时间 秒
WAIT_TIME = 20

# 时间区间内最大数据量
MAX_DATA_COUNT = 5000

# 抓取多少页后重启浏览器
MAX_CRAWL_PAGE_COUNT = 50

# 时间格式format
DATETIME_FORMAT = "%Y-%m-%d %H:%M:%S"

DATA_DIR = "../data/"

class redconfig():
    def __init__(self,filepath=None):
        if filepath:
            self.config_path=filepath
        else:
            self.config_path=os.path.join(os.path.dirname(os.getcwd()), 'config.ini')
        print(self.config_path)
        self.config=configparser.ConfigParser()
        self.config.read(self.config_path,encoding='utf-8')

    def getValueByDict(self,key,value):
        return self.config.get(key,value)

    def getDictBySection(self,section):
        return dict(self.config.items(section))


def print_it():
    print("kkk")

if __name__ == '__main__':
    myconfig=redconfig()
    # print(myconfig.getValueByDict('yqt_info','username'))
    # print(myconfig.getDictBySection('time_info'))
    # cron_info=myconfig.getDictBySection('cron_info')
    # # for key,value in cron_info.items():
    # #     cron_info[key]=eval(value)
    # tigger1=CronTrigger(**cron_info)
    # print(tigger1.fields)
    # scheduler=BlockingScheduler()
    # scheduler.add_job(print_it,tigger1,max_instances=10,id='212')
    # scheduler.start()
    # test1=eval(myconfig.getValueByDict("crawl_condition", "condition"))
    # print(type(test1))
    # print(test1)
    # for key,value in test1.items():
    #     print(key,value)
    # dir_path = os.path.dirname(os.path.dirname(os.path.relpath(__file__)))
    # dir_path=os.path.dirname(os.getcwd())
    # config_xlsx_path = os.path.join(dir_path, r'config.xlsx')
    # wb = load_workbook(config_xlsx_path, data_only=True)

    test1 = eval(myconfig.getValueByDict("industry_info", "project_name"))
    print(test1)
    print(type(test1))
    # for i in test1:
    print('优速测试_快消品' in test1)

