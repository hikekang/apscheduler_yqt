# -*- coding=utf-8 -*-
import time
import datetime
import os

from apscheduler.schedulers.blocking import BlockingScheduler
from apscheduler.triggers.cron import CronTrigger
from pyquery import PyQuery as pq
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.action_chains import ActionChains

from utils.mylogger import logger
from utils.spider_helper import SpiderHelper
from utils.webdriverhelper import MyWebDriver
from utils.my_pyautogui import pyautogui
from utils.webdriverhelper import WebDriverHelper
from yuqingtong import config
import requests
from openpyxl import load_workbook
from utils import ssql_helper
class YQTSpider(object):

    def __init__(self,info,spider_driver=None, start_time=None,end_time=None,  *args, **kwargs):
        if spider_driver is None:
            self.spider_driver = WebDriverHelper.init_webdriver(is_headless=config.HEAD_LESS)  # type:MyWebDriver
        else:
            self.spider_driver = spider_driver  # type:MyWebDriver
        self.wait = WebDriverWait(self.spider_driver, config.WAIT_TIME)
        self.info=info
        # 需要替换
        self.data_file_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), f"data\{info['project_name']}\{datetime.datetime.now().strftime('%Y-%m-%d')}",
                                           f"{self}_{info['yuqingtong_username']}_{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}_{start_time}_{end_time}.xlsx".replace(':','_'))
        self.process_file_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                                              "process.txt")


        # 上次终止时间
        self.last_end_time = None  # type:datetime.datetime

        # 本次终止时间
        self.next_end_time = None  # type:datetime.datetime

        # self.last_end_time = self.end_time
        default_start_time = datetime.datetime.combine(
            (datetime.datetime.now() + datetime.timedelta(days=-99)).date(),
            datetime.datetime.min.time())
        default_end_time = datetime.datetime.combine(datetime.datetime.now().date(), datetime.datetime.min.time())
        print(default_end_time)
        self.interval = [default_start_time, default_end_time]
        self.next_page_num = 1

        self.crawl_page_count = 0  # 计数用，重启浏览器使用

    def __str__(self):
        return "舆情通"


    def _login(self, count=1):
        """
        登录
        :return: True OR False
        """
        driver = self.spider_driver
        if count > 10:
            logger.warning("登录次数超过10次,请检查登录流程")
            return False

        logger.info(f"开始登录....{count}")
        driver.get("http://yuqing.sina.com/staticweb/#/login/login")
        username = driver.find_element_by_xpath("//input[@formcontrolname='userName']")
        password = driver.find_element_by_xpath("//input[@formcontrolname='password']")
        yqzcode = driver.find_element_by_xpath("//input[@formcontrolname='yqzcode']")

        submit_buttion = driver.find_element_by_xpath("//button[contains(@class,'login-form-button')]")

        username.send_keys(self.info['yuqingtong_username'])
        password.send_keys(self.info['yuqingtong_password'])

        logger.info("获取验证码....")
        while 1:
            code_img = None
            try:
                code_img = self.wait.until(
                    EC.presence_of_element_located((By.XPATH, "//img[contains(@src,'validate/image')]")))
            except TimeoutException:
                pass

            if code_img and code_img.size.get("width") > 0:
                break
            logger.info("验证码图片没加载出来，刷新...")
            driver.refresh()
        code_img_base64 = code_img.screenshot_as_base64
        code = SpiderHelper.recognise_code(code_img_base64,self.info)
        logger.info(f"获取验证码:{code}")
        if not code:
            code = "1234hi"

        yqzcode.send_keys(code)
        submit_buttion.click()

        try:
            wait = self.wait.until(
                EC.invisibility_of_element_located((By.XPATH, "//input[@formcontrolname='userName']")))
            print(wait)
            print("登录成功")
            return True
        except Exception as e:
            logger.warning(e)

        try:
            mobile_code = driver.find_element_by_css_selector(".ant-modal-content .mt20 input")
            # pyautogui.prompt("需要手机验证")
            while not driver.is_url_change():
                print("还没填写验证码")
                time.sleep(1)

        except NoSuchElementException:
            pass

        if driver.is_url_change():
            print("登录成功")
            return True
        else:
            return self._login(count=count + 1)
    # 解析页面进行数据抓取和保存
    def _parse(self, page_source):
        doc = pq(page_source)

        items = doc.find('tbody tr.ng-scope').items()

        data_list = []
        for item in items:
            td_title = item.find('td:nth-child(2)')
            td_origins = item.find('td:nth-child(4)')
            td_time = item.find('td:last-child')

            # print(td_time)

            def parse_time(td_time):
                from datetime import datetime
                ymd_1 = td_time.find('span:first-child').text()
                ymd_2 = td_time.find('span:last-child').text()
                try:
                    if "年" in ymd_1:  # 不是今年的数据
                        ymd = "".join([ymd_1, ymd_2])
                    elif "今天" in ymd_2:  # 今天的数据
                        ymd = " ".join([f"{str(datetime.now().date())}", ymd_1])
                        return datetime.strptime(ymd, "%Y-%m-%d %H:%M").strftime("%Y-%m-%d %H:%M:%S")
                    else:
                        ymd = " ".join([f"{datetime.now().year}年" + ymd_2, ymd_1])
                    return datetime.strptime(ymd, "%Y年%m月%d日 %H:%M").strftime("%Y-%m-%d %H:%M:%S")
                except Exception as e:
                    logger.warning(e)
                    logger.warning(td_time)

            # 转发类型
            sort = td_title.find('div.news-item-tools.font-size-0 span.tag-yuan').text()

            repost_content = td_title.find('div.item-title.resend-news-item-title').text().replace('\n', '')
            pinglun=td_title.find('div.inline-block.mr10.tag-ping').text().replace('\n','')#评论

            if repost_content:
                sort = "转发"
            if pinglun:
                sort = "评论"

            content = td_title.find('div.item-title.news-item-title.contenttext.ng-binding').text().replace('\n', ''),

            title = td_title.find('span[ng-bind-html="icc.title | trustAsHtml"]').text().replace('\n', '')
            site_name = td_origins.find('span[ng-bind="icc.captureWebsiteName"]').text()

            if (site_name == '新浪微博' or site_name=='今日头条'):
                data = {
                    '时间': parse_time(td_time),
                    '标题': content[0],
                    '描述': '',
                    '链接': td_title.find('.news-item-tools.font-size-0 div:nth-child(2)>div>ul>li:nth-child(4) a').attr(
                        "href"),
                    '转发内容': repost_content,
                    '发布人': td_title.find('a[ng-bind="icc.author"]').text(),
                    'attitude': td_title.find(
                        'div[ng-show="view.resultPresent != 3"] div.sensitive-status-content:not(.ng-hide)>span:first-child').text(),
                    'images': ",".join([img.attr('src') for img in item.find('.actizPicShow img').items()]),
                    'reposts_count': item.find(
                        'div.news-item-title.color-gray-6.font-size-12.ng-scope span:first>span').text(),

                    'comments_count': item.find(
                        'div.news-item-title.color-gray-6.font-size-12.ng-scope span:nth-child(3)>span').text(),

                    'sort': sort,

                    'industry': td_title.find(
                        '.news-item-tools.font-size-0 div:nth-child(1) div[ng-if="secondTrade!=null"]').text(),

                    'related_words': td_title.find(
                        '.news-item-tools.font-size-0 div:nth-child(1) span[ng-bind="icc.referenceKeyword"]').text(),

                    'site_name': td_origins.find('span[ng-bind="icc.captureWebsiteName"]').text(),
                    'area': td_origins.find('div[ng-bind="icc.province"]').text(),

                }
                if sort=='评论':
                    data['链接']=td_title.find('.news-item-tools.font-size-0 div:nth-child(2)>div>ul>li:nth-child(3) a').attr("href")
                    biaoti=td_title.find('div.item-title.news-item-title.dot.ng-binding').text().replace('\n', '')
                    data['标题']=biaoti

            else:
                data = {
                    '时间': parse_time(td_time),
                    '标题': title,
                    '描述': content[0],
                    '链接': td_title.find('.news-item-tools.font-size-0 div:nth-child(2)>div>ul>li:nth-child(4) a').attr(
                        "href"),
                    '转发内容': repost_content,
                    '发布人': td_title.find('a[ng-bind="icc.author"]').text(),
                    'attitude': td_title.find(
                        'div[ng-show="view.resultPresent != 3"] div.sensitive-status-content:not(.ng-hide)>span:first-child').text(),
                    'images': ",".join([img.attr('src') for img in item.find('.actizPicShow img').items()]),
                    'reposts_count': item.find(
                        'div.news-item-title.color-gray-6.font-size-12.ng-scope span:first>span').text(),

                    'comments_count': item.find(
                        'div.news-item-title.color-gray-6.font-size-12.ng-scope span:nth-child(3)>span').text(),

                    'sort': sort,

                    'industry': td_title.find(
                        '.news-item-tools.font-size-0 div:nth-child(1) div[ng-if="secondTrade!=null"]').text(),

                    'related_words': td_title.find(
                        '.news-item-tools.font-size-0 div:nth-child(1) span[ng-bind="icc.referenceKeyword"]').text(),

                    'site_name': td_origins.find('span[ng-bind="icc.captureWebsiteName"]').text(),
                    'area': td_origins.find('div[ng-bind="icc.province"]').text(),

                }
            # print(data)
            data_list.append(data)
        return data_list

    def parse_data(self):
        """
        解析页面生成数据
        :return: 数据已字典形式存在list中
        """
        driver = self.spider_driver
        page_source = driver.page_source
        return self._parse(page_source)

    def _set_conditions(self, start_time, end_time):
        """
        设置筛选条件
        :param start_time: 起始时间
        :param end_time: 终止时间
        :return: True or Flase
        """
        driver = self.spider_driver
        print(start_time)
        start_time_str = start_time.strftime(config.DATETIME_FORMAT)
        end_time_str = end_time.strftime(config.DATETIME_FORMAT)
        driver.find_element_by_css_selector('div.inline-block.custom-time-period').click()

        start_input = driver.find_element_by_css_selector("#startTimeInput1")
        start_input.clear()
        start_input.send_keys(start_time_str)
        time.sleep(0.2)
        end_input = driver.find_element_by_css_selector("#endTimeInput1")
        end_input.clear()
        end_input.send_keys(end_time_str)
        time.sleep(0.2)
        driver.find_element_by_css_selector('span[ng-click="confirmTime(1)"]').click()
        time.sleep(0.2)
        # -----------点击全部------------------
        driver.find_element_by_css_selector('#informationContentType0').click()
        time.sleep(0.2)
        driver.find_element_by_css_selector('#select0').click()
        time.sleep(0.2)
        # ------------------------------------
        driver.find_element_by_css_selector("#searchListButton").click()
        # time.sleep(2)
        return True

    # 细化时间，分开爬取
    def _adapt_time_interval(self):
        """
        改变时间的区间
        :param start_time:
        :param end_time:
        :return: 截止时间
        """
        # end_time = datetime.datetime.now()
        # start_time = end_time + datetime.timedelta(days=-30)

        while 1:
            self.spider_driver.scroll_to_top()
            logger.info("设置时间区间...")
            self._save_process()
            # self._set_conditions(start_time, end_time)
            """
                假如多项目切换，提前换词
            """
            # 设置时间
            self._set_conditions(self.last_end_time, self.next_end_time)


            if not self._is_page_loaded():
                logger.info("设置时间时页面加载出现问题")
                return False
            if not self._is_data_count_outside():  # 没有超过5000条，不用调整
                logger.info(f"小于{config.MAX_DATA_COUNT}条,符合条件")
                break
            logger.info(f"页面数据大于{config.MAX_DATA_COUNT}条，调整时间区段")

            #时间设置
            timedelta = self.next_end_time - self.last_end_time
            if timedelta.days <= 1:
                timedelta_hours = timedelta.total_seconds() / 60 / 60
                if timedelta_hours <= 1:
                    logger.info("时间区间小于一小时，无法调整")
                    break
                logger.info("时间区间小于一天，继续细化调整")
                # 下次开始时间
                self.next_end_time = self.last_end_time + datetime.timedelta(hours=timedelta_hours / 2)
            else:
                # return None
                self.next_end_time = datetime.datetime.combine(self.last_end_time.date(),
                                                               datetime.datetime.min.time()) + \
                                     datetime.timedelta(days=round(timedelta.days / 2))
            # if timedelta.days % 2 == 0:  # 偶数
            #     end_time = start_time.date() + datetime.timedelta(days=timedelta.days / 2)
            # else:
            #     end_time = start_time.date() + datetime.timedelta(days=(timedelta.days + 1) / 2)

        logger.info(
            f"当前时间区间:{self.last_end_time.strftime(config.DATETIME_FORMAT)}  --"
            f" {self.next_end_time.strftime(config.DATETIME_FORMAT)}")

    @property
    def _maxpage(self):
        # 获取最大页数
        return int(self.spider_driver.find_element_by_css_selector('span.page-number-total.ng-binding').text)

    @property
    def _count_number(self):
        return int(self.spider_driver.find_element_by_css_selector('span[ng-bind="originStat.total"]').text)
    def _is_data_count_outside(self):
        """
        数据量是否超出5000
        :return:
        """
        try:
            data_count = int(self.spider_driver.find_element_by_css_selector(
                'span[ng-bind="originStat.total"]').text)
            logger.info(f"当前数据量:{data_count}")
            if data_count > config.MAX_DATA_COUNT:
                return True
        except Exception as e:
            logger.warning(e)
            return True

    def _is_page_loaded(self, count=1):
        """
        判断页面时候是否加载完全
        :return:
        """
        if count > 3:
            logger.warning("页面加载可能卡住")
            return False
        try:
            # 加载数据时的圆圈消失就说明页面加载完全
            # logger.debug(f"检测是否加载完全。。{count}")
            # loading_ico = self.spider_driver.find_element_by_css_selector('#iccListLoading .spinner').is_displayed()
            # print(loading_ico)
            wait_data_count_loading_disappear = self.wait.until(
                EC.invisibility_of_element_located((By.CSS_SELECTOR, '#originStatId>.spinner')))

            wait_loading_disappear = self.wait.until(
                EC.invisibility_of_element_located((By.CSS_SELECTOR, '#iccListLoading>.spinner')))
            if wait_loading_disappear:
                try:
                    wait_tr_appear = self.wait.until(
                        EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'tbody.contenttext>tr')))
                except TimeoutException:
                    try:
                        # 本身就没数据显示没数据
                        wait_no_data_appear = self.wait.until(
                            EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'p.noData-word')))
                    except TimeoutException:
                        return False

            logger.info("页面加载完全")
            return True

        except TimeoutException as e:
            logger.warning(e)
            print('再次判断')
            return self._is_page_loaded(count=count + 1)


    def _switch_data_count_perpage(self):
        """
        点击每页100条按钮
        :return:
        """
        print('更改100')
        if not self._is_page_loaded():
            logger.info("更改100条数据/每页前，页面加载有问题")
            return False
        print('开始更改')
        self.spider_driver.find_element_by_css_selector('div[ng-show="selectType == 1"]').click()
        print('点击')
        time.sleep(1)
        print('查找sleep')
        self.spider_driver.find_element_by_css_selector(
            'div[ng-show="selectType == 1"] li[data-original-index="2"]>a').click()
        print('选择100')
        if not self._is_page_loaded():
            logger.info("更改100条数据/每页后，页面加载有问题")
            return False
        logger.info("更改100条数据/每页")
        print('更改完毕')
        return True

    def _get_input_time_range(self):
        """
        获取输入的时间区间
        :return:
        """
        while 1:
            input_start_time = pyautogui.prompt('请输入起始日期\n格式：2020-01-01', "起始时间",
                                                default=self.interval[0].strftime('%Y-%m-%d'))
            # input_start_time = config.info['start_time'].strftime('%Y-%m-%d')
            try:
                start_time = datetime.datetime.strptime(input_start_time, "%Y-%m-%d")
                break
            except ValueError as e:
                logger.warning(e)
                pyautogui.alert(f"检查日期是否有误：{input_start_time}")
            except TypeError:
                return
        while 1:
            input_end_time = pyautogui.prompt(f'起始时间：{input_start_time}\n请输入终止日期\n格式：2020-12-01', "终止时间",
                                              default=self.interval[1].strftime('%Y-%m-%d'))
            # input_end_time = config.info['end_time'].strftime('%Y-%m-%d')
            try:
                end_time = datetime.datetime.strptime(input_end_time, "%Y-%m-%d") + datetime.timedelta(
                    days=1)
                if end_time <= start_time:
                    pyautogui.alert(f"终止时间[{input_end_time}]小于起始时间[{input_start_time}]")
                    continue

                if (end_time - start_time).days > 100:
                    pyautogui.alert(f"时间跨度不能超过100天，请重新输入")
                    return self._get_input_time_range()
                break
            except ValueError as e:
                logger.warning(e)
                pyautogui.alert(f"检查日期是否有误：{input_start_time}")
            except TypeError:
                return
        return start_time, end_time

    def _reload(self):
        """
        刷新页面
        :return:
        """
        self.spider_driver.refresh()
        self.spider_driver.refresh()
        if not self._is_page_loaded():
            return False
        return True

    def _go_page_num_by_conditions(self, is_reload=False):
        """
        根据条件和页数进入某一特定页面
        如：2020-01-01 00:00:00 至2020-01-02 00:00:00 第10页
        :return:
        """

        if is_reload:
            logger.info(f'重新进入此页面：{self.last_end_time.strftime(config.DATETIME_FORMAT)} '
                        f'- {self.next_end_time.strftime(config.DATETIME_FORMAT)} '
                        f'第{self.next_page_num}页')
            self._reload()
        # else:
        # logger.info(f'进入此页面：{self.condition.get("start_time").strftime(config.DATETIME_FORMAT)} '
        #             f'- {self.condition.get("end_time").strftime(config.DATETIME_FORMAT)} ')

        self._adapt_time_interval()

        if self.next_page_num > 1:
            logger.info(f"直接进入第{self.next_page_num}页")
            self.spider_driver.find_element_by_css_selector('input[ng-model="goPage"]').send_keys(
                self.next_page_num)
            time.sleep(0.5)
            self.spider_driver.find_element_by_css_selector('div[ng-click="gotoSearchList();"]').click()
        if not self._is_page_loaded():
            logger.info("直接进入第多少页时页面加载出现问题")
            return False
        page_num = self.spider_driver.find_element_by_css_selector("#listPage").text
        if int(page_num) != self.next_page_num:
            logger.warning("页面上当前页面和应该进入的页面不一样，请检查")
            return False
        logger.info("进入成功")
        return True

    def _save_process(self):
        with open(self.process_file_path, "w", encoding="utf-8") as f:
            f.write(f"{self.interval[0]}至{self.interval[1]}|"
                    f"{self.last_end_time}_{self.next_end_time}_{self.next_page_num}_{self.data_file_path}")
    def _turn_page(self, max_page_num,time_sleep):
        """
        翻页
        :return:
        """

        # 翻页开始
        while 1:
            if self.crawl_page_count > config.MAX_CRAWL_PAGE_COUNT:
                self.crawl_page_count = 0
                return "restart_browser"
            # todo:每解析完一页数据，就保存下当前任务进程到文件中：时间区间，第几页,保存文件路径
            # self._save_process()
            if self.next_page_num > 1:
                if not self._is_page_loaded():
                    return False
            # print("页面加载完成")

            logger.info(f"当前第【{self.next_page_num}】页,共{max_page_num}页")
            data_list = self.parse_data()
            logger.info(f"解析到{len(data_list)}条数据")
            SpiderHelper.save_xlsx(data_list=data_list, out_file=self.data_file_path)
            logger.info(f"保存完毕")

            if self.next_page_num >= max_page_num:
                logger.info("抓取到最大页，停止")
                data_count = int(self.spider_driver.find_element_by_css_selector(
                    'span[ng-bind="originStat.total"]').text)

                break
            self.next_page_num += 1
            logger.info(f"点击下一页.....")
            self._save_process()
            try:
                self.spider_driver.find_element_by_xpath('//i[@class="fa-page-arrow-right ng-scope"]').click()
            except NoSuchElementException:
                logger.info("没有找到下一页的按钮")
                break
            self.crawl_page_count += 1

        time.sleep(time_sleep)
        return True

    def _crawl2(self,time_sleep):

        logger.info(
            f'任务时间区间：{self.interval[0].strftime(config.DATETIME_FORMAT)} --- '
            f'{self.interval[1].strftime(config.DATETIME_FORMAT)}')

        # 改成每页100条
        self._switch_data_count_perpage()

        is_reload = False
        set_conditions_reload_count = 0
        turn_page_reload_count = 0

        while 1:
            # 设置时间区间
            if not self._go_page_num_by_conditions(is_reload):
                if set_conditions_reload_count >= 3:
                    logger.warning("设置时间区间时，连续出现问题3次，退出")
                    return False
                set_conditions_reload_count += 1
                is_reload = True
                continue
            set_conditions_reload_count = 0
            max_page_num = self._maxpage
            print(f"最大页数：{max_page_num}")

            # 翻页并抓取数据
            resp = self._turn_page(max_page_num,time_sleep)
            if resp == "restart_browser":
                return resp
            elif not resp:
                self._turn_page(max_page_num, time_sleep)
                if turn_page_reload_count >= 3:
                    logger.warning("翻页时，连续出现问题3次，退出")
                    return False
                turn_page_reload_count += 1
                is_reload = True
                continue
            turn_page_reload_count = 0
            # 翻页结束

            # 设置下次抓取条件
            self.next_page_num = 1
            if self.next_end_time >= self.interval[1]:
                logger.info("解析到终止时间，抓取完成")
                logger.info("上传数据")
                post_url="http://localhost:8086/localproject/industry/industryBigDataExcelByEasyExcel"
                files={'file':open(self.data_file_path,'rb')}
                proxies = {"http": None, "https": None}
                post_info=requests.post(post_url,files=files,proxies=proxies).text
                post_info = eval(post_info)
                logger.info("开始记录")
                #舆情通数量
                yqt_count=self._count_number
                #xlsx数据量
                wb=load_workbook(self.data_file_path)
                xlsx_num=wb[wb.sheetnames[0]].max_row
                record_file_path=os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                             f"record\{config.info['project_name']}",f"{self}_记录.xlsx")
                sql_number=ssql_helper.find_info_count(self.interval[0],self.interval[1],config.info['sheet_name'])
                data_list=[config.info['project_name'],datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),self.last_end_time,self.next_end_time]
                SpiderHelper.save_record(record_file_path,yqt_count,xlsx_num,post_info['number'],sql_number,data_list=data_list)
                return True
            else:
                self.last_end_time = self.next_end_time  # 上次终止时间就是下次起始时间
                self.next_end_time = self.interval[1]

    def _load_condition_process_file(self):
        if os.path.exists(self.process_file_path):
            print(self.process_file_path)
            with open(self.process_file_path, "r", encoding="utf-8") as f:
                last_task_process = f.readline()
            return last_task_process
    # 关键词修改
    def modifi_keywords(self,keywords):
        """
        读取config。xlsx文件关键词
        根据关键词进行修改
        """
        driver = self.spider_driver
        li = driver.find_element_by_xpath('//li[@class="site-menu-item is-shown open"]')
        span = li.find_element_by_xpath('//span[@class="fa-tree-plan-tools-bar"]')
        action = ActionChains(driver)
        action.move_to_element(li).perform()
        span.click()
        driver.find_element_by_xpath('//li[@class="add-plan-trigger"]/a').click()
        keywords = driver.find_element_by_xpath('//div[@class="edit_textarea mb5 ng-binding"]')
        keywords.clear()
        keywords.send_keys(keywords)
        # 保存
        driver.find_element_by_id("saveHighSetKeyword").click()
        time.sleep(1)

    def start(self,start_time,end_time,time_sleep):
        try:
            # 1.登录
            if not self._login():
                raise Exception("登录环节出现问题")
            self.interval = [start_time, end_time]
            self.last_end_time = self.interval[0]
            self.next_end_time = self.interval[1]
            # 抓取数据
            resp = self._crawl2(time_sleep)
            if resp == "restart_browser":
                logger.info("重启浏览器")
                self.spider_driver.quit()
                self.spider_driver = WebDriverHelper.init_webdriver(is_headless=config.HEAD_LESS)
                self.wait = WebDriverWait(self.spider_driver, config.WAIT_TIME)
                self.start(resp)
            elif resp is True:
                os.remove(self.process_file_path)
                # pyautogui.alert("抓取完成...")
                os.system(f'explorer /select,{self.data_file_path}')
        except Exception as e:
            # logger.warning(e)
            print(e)
        finally:
            if self.spider_driver.service.is_connectable():
                self.spider_driver.quit()


#修改xlsx文件进行自动抓取
def xlsx_work():
    time_list = config.get_time_list()
    # print(time_list)
    infos = config.row_list
    for info in infos:
        for tim in time_list:
            print('抓取次数')
            print(tim['start_time'])
            yqt_spider = YQTSpider(info,start_time=tim['start_time'], end_time=tim['end_time'])
            yqt_spider.start(start_time=tim['start_time'], end_time=tim['end_time'], time_sleep=tim['time_delay'])
            config.info['start_time'] = tim['start_time']
            config.info['end_time'] = tim['end_time']
# 自定义时间抓取任务
def work_it():
    end_time = datetime.datetime.now().strftime('%Y-%m-%d %H') + ":00:00"
    one_hour_ago = datetime.datetime.now() - datetime.timedelta(hours=1)

    start_time = one_hour_ago.strftime('%Y-%m-%d %H') + ":00:00"
    # 开始时间
    start_time = datetime.datetime.strptime(start_time, "%Y-%m-%d %H:%M:%S")
    # 结束时间
    end_time = datetime.datetime.strptime(end_time, "%Y-%m-%d %H:%M:%S")
    infos=config.row_list
    for info in infos:
        yqt_spider = YQTSpider(info,start_time=start_time, end_time=end_time)
        yqt_spider.start(start_time=start_time, end_time=end_time, time_sleep=2)
    #
def apscheduler():
    trigger1 = CronTrigger(hour='9-18', minute='*/5', second=0, jitter=30)
    sched = BlockingScheduler()
    sched.add_job(work_it, trigger1, id='my_job_id')
    sched.start()

if __name__ == '__main__':
    # apscheduler()
    xlsx_work()