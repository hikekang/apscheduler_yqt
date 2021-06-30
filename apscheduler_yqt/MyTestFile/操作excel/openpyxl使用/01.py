#!/usr/bin/python3.x
# -*- coding=utf-8 -*-
"""
 Time       : 2021/6/18 10:37
 Author     : hike
 Email      : hikehaidong@gmail.com
 File Name  : 01.py
 Description:
 Software   : PyCharm
"""
from openpyxl.workbook import Workbook
# 不需要创建文件
wb=Workbook()

ws=wb.active

#默认插入到最后
ws1=wb.create_sheet("test1")

#插入到最前
ws2=wb.create_sheet("test2",0)

#插入到倒数第二
ws3=wb.create_sheet("test3",-1)


#修改sheet名称
ws.title="hike"

# 单元格赋值
ws['A4']=11

#单元格值获取
A_4=ws['A4']

# 通过行号和列号访问单元格
ss=ws.cell(row=1,column=1)

ss=ws.cell(row=1,column=1,value="赋值")

#访问多个单元格
# cell_range=ws['A1','C1']

#访问列或行
col_c=ws['C']

row=ws[10]

for row in ws.iter_rows(min_row=1,min_col=3,max_row=2):
    for cell in row:
        print(cell)

# 查看sheet名称
print(wb.sheetnames)
for sheet in wb:
    print(sheet.title)
    print(sheet.values)

# wb.save("test.xlsx")