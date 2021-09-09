# -*- coding: utf-8 -*-
"""
   File Name：     test_openpyxl
   Description :
   Author :       hike
   time：          2021/3/31 10:00
"""
from openpyxl import Workbook
from openpyxl import load_workbook
import os

def save_xlsx(data_list, out_file):
    dir = os.path.dirname(out_file)
    if not os.path.exists(dir):
        os.mkdir(dir)
    wb = load_workbook(out_file)
    sheet = wb['Sheet1']
    # 写入表头
    head_xlsx = ['时间', '标题', '描述', '链接', '转发内容', '发布人', 'attitude', 'images', 'reposts_count', 'comments_count', 'sort',
                 'industry', 'related_words', 'site_name', 'area']
    for i, item in enumerate(head_xlsx):
        sheet.cell(row=1, column=i + 1, value=item)
    for i, item in data_list:
        sheet.cell(row=i + 2, column=i + 1, value=item)
    pass
# save_xlsx()
datalisy=[]
for i in range(10):
    data={

    }


import openpyxl

products = [{'id':46329,
             'discription':'AD BLeu',
             'marque':'AZERT',
             'category':'liquid',
             'family': 'ADBLEU',
             'photos':'D:\\hamzawi\\hamza\\image2py\\46329_1.png'}
            ]

# Dictionarys are not in order by default
# Define a `list` of `keys` in desired order
fieldnames = ['id', 'discription', 'marque', 'category', 'family', 'photos']

# create a new workbook
wb = openpyxl.load_workbook('11.xlsx')
ws = wb['Sheet1']

# append headers
ws.append(["Product ID", "Product Name", "Marque", "Category", "Family", "Photos"])

# append data
# iterate `list` of `dict`
for product in products:
    # create a `generator` yield product `value`
    # use the fieldnames in desired order as `key`
    values = (product[k] for k in fieldnames)

    # append the `generator values`
    ws.append(values)

# show Worksheet Values
for row_values in ws.iter_rows(values_only=True):
    for value in row_values:
        print(value, end='\t')
    print()
wb.save('11.xlsx')
# wb.save()