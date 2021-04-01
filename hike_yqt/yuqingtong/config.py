# -*- coding: utf-8 -*-



from openpyxl import load_workbook

#数据加载

wb=load_workbook('../config.xlsx',data_only=True)
sheet=wb['Sheet1']
row_info=[]
for row in sheet.iter_rows(min_row=2,max_col=7,max_row=2):
    for cell in row:
        row_info.append(cell.value)

pinyin=['yuqingtong_username','yuqingtong_password','dama_username','dama_password','softid','project_name','sheet_name']
info=dict(list(zip(pinyin,row_info)))

USERNAME = info['yuqingtong_username']
PASSWORD = info['yuqingtong_password']

time_data_all=sheet.iter_rows(min_row=7)

def get_time_list():
    times = []
    for row in time_data_all:
        time = {
            'start_time':'',
            'end_time':'',
            'time_delay':'2'
        }
        for cell in row:
            if(cell.column==1):
                time['start_time']=cell.value
            elif(cell.column==2):
                time['end_time']=cell.value
            elif(cell.column==3):
                time['time_delay']=cell.value
        times.append(time)
    return times


USERNAME = info['yuqingtong_username']
PASSWORD = info['yuqingtong_password']
# 浏览器相关
HEAD_LESS = False

# 等待时间 秒
WAIT_TIME = 20

# 时间区间内最大数据量
MAX_DATA_COUNT = 5000

# 抓取多少页后重启浏览器
MAX_CRAWL_PAGE_COUNT = 60

# 时间格式format
DATETIME_FORMAT = "%Y-%m-%d %H:%M:%S"

DATA_DIR = "../data/"
