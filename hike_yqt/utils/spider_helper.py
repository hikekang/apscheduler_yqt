import requests
import json
import csv
import pandas as pd
import os

from openpyxl import Workbook

from yuqingtong.config import *


class SpiderHelper(object):

    @staticmethod
    def recognise_code(image_base64):
        # base64_data = base64.b64encode(img_bytes)
        # img_b64_data = base64_data.decode()
        data = {
            "user": info['dama_username'],
            "pass": info['dama_password'],
            "softid": info['softid'],
            "codetype": 1006,
            "file_base64": image_base64
        }
        headers = {"Content-Type": "application/json"}
        try:
            proxies = {"http": None, "https": None}
            response = requests.post(url="http://upload.chaojiying.net/Upload/Processing.php", data=json.dumps(data),
                                     headers=headers, proxies=proxies)
            print(response.json())
            return response.json().get("pic_str")
        except Exception as e:

            print(e)

    @staticmethod
    def save_csv(data_list, out_file):
        # print(out_file)
        dir = os.path.dirname(out_file)
        # print(dir)
        if not os.path.exists(dir):
            os.mkdir(dir)
        # 利用datafram 获取列索引
        df = pd.DataFrame(data_list)
        # print(list(df.columns))

        # 通过判断时候有csv文件，来判断是否写表头，避免重复写表头
        is_writeheader = not os.path.exists(out_file)
        with open(out_file, 'a', encoding="utf-8-sig", newline="") as csvfile:

            writer = csv.DictWriter(csvfile, fieldnames=df.columns)
            if is_writeheader:
                writer.writeheader()  # 写入表头

            for data in data_list:
                writer.writerow(data)

    @staticmethod
    def save_xlsx(data_list, out_file):
        head_xlsx = ['时间', '标题', '描述', '链接', '转发内容', '发布人', 'attitude', 'images', 'reposts_count', 'comments_count',
                     'sort', 'industry', 'related_words', 'site_name', 'area']

        dir = os.path.dirname(out_file)
        if not os.path.exists(dir):
            os.makedirs(dir)
        if not os.path.exists(out_file):
            wb = Workbook(out_file)
            sheet = wb.create_sheet(title=info['sheet_name'])
            # 写入表头
            # for i,item in enumerate(head_xlsx):
            #     sheet.cell(row=1,column=i+1,value=item)
            sheet.append(head_xlsx)
            for data in data_list:
                values = (data[k] for k in head_xlsx)
                sheet.append(values)
            wb.save(out_file)
            wb.close()
        else:
            wb = load_workbook(out_file)
            sheet = wb[info['sheet_name']]
            for data in data_list:
                values = (data[k] for k in head_xlsx)
                sheet.append(values)
            wb.save(out_file)
            wb.close()

if __name__ == '__main__':
    SpiderHelper.save_csv([{"name": 12}], os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "resle.csv"))
    with open("./image.png", "rb") as f:
        img_bytes = f.read()
    print(SpiderHelper.recognise_code(img_bytes))
